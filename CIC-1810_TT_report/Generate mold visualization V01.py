import os
from openpyxl import load_workbook, Workbook
from PIL import Image, ImageDraw, ImageFont
from datetime import datetime

# Script for visualisation of measurement data of wall thickness on RIC_1881_2 closures

# What this script does:
# 1. Find excel file with measurement data (script and excel file must be in the same folder)
# 2. Create a list that contains relevant measurement data: [Cavity_diameter, Core_diameter, Thickness_12, Thickness_3, Thickness_9]
# 3. Assuming, that position of Cavity_diameter is (0, 0), use measurement data to calculate position of Core_diameter
# 4. Plot this information in intuitive way to image file to be used in excel for reporting

# Define global scale (make image bigger or smaller)
global_scale = 10

# Define all functions
def excel_files_in_directory():
    
    # Get the directory where the Python file is saved
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # List all files in that directory
    all_files = os.listdir(script_dir)

    # Make an empty list for excel files
    files_to_return = []

    #Filter files to leave only .xlsx
    for file in all_files:
        if file[-5:] == '.xlsx':
            files_to_return.append(file)

    # Return excel file that contains measurement data
    if len(files_to_return) > 1:
        print('There are several excel files in this folder!')
        return('There are several excel files in this folder!')
        
    if len(files_to_return) == 0:
        print('There are no excel files in this folder!')
        return('There are no excel files in this folder!')

    return(files_to_return[0])

def data_list(file, start, stop):
    """
    Extract relevant measurement data from excel file and add it to a list
    Excel file and script fyle must be in the same directory
    Tab with measurement data must be active in excel file
    Data from following collumns will be used: 'Outer DIA', 'Inner DIA', 'Measured wall thickness @12, @3, @9'
    """

    # start, stop = first row with measurement data in excel, last row with measurement data in excel
    # collumns = list with collumn names that contain relevant data (collumn names in string format)

    #     'Outer DIA', 'Inner DIA', 'Thickness 12', 'Thickness 3', 'Thickness 9'
    collumns = [  'C',         'D',            'Q',           'H',           'N']

    wb = load_workbook(filename = data_file)
    count = start
    result = []

    while count < stop + 1:
        outer_diameter = wb.active[collumns[0] + str(count)].value
        inner_diameter = wb.active[collumns[1] + str(count)].value
        thickness_12   = wb.active[collumns[2] + str(count)].value
        thickness_3    = wb.active[collumns[3] + str(count)].value
        thickness_9    = wb.active[collumns[4] + str(count)].value

        result.append([outer_diameter, inner_diameter, thickness_12, thickness_3, thickness_9])
        count = count + 1

    return(result)

def get_range():
    # From number of entries in excel file calculate number of cavities,
    # Find first and last row with data to read

    data_file = excel_files_in_directory()
    position = 20                           # First row with data is always 'B20'
    count = 0                               # How many cavities?
    condition = True

    wb = load_workbook(filename = data_file)

    while condition == True:
        cell_value = wb.active['B' + str(position)].value
        if type(cell_value) != type(wb.active['B20'].value):
            condition = False
        else:
            position = position + 1
            count = count + 1

    # Return: ('First row where to start'), ('Last row, or where to finish'), ('Number of cavities')
    return(20, position - 1, count)

def calculate_circle_center(A, B, C):
    # From 3 points with coordinates (x, y), calculate center point of circle
    
    yDelta_a = B[1] - A[1]
    xDelta_a = B[0] - A[0]
    yDelta_b = C[1] - B[1]
    xDelta_b = C[0] - B[0]

    aSlope = yDelta_a / xDelta_a
    bSlope = yDelta_b / xDelta_b

    center_x = (aSlope * bSlope * (A[1] - C[1]) + bSlope * (A[0] + B[0]) - aSlope * (B[0] + C[0])) / (2 * (bSlope - aSlope))
    center_y = (-1 * (center_x - (A[0] + B[0]) / 2) / aSlope + (A[1] + B[1]) / 2)

    return center_x, center_y

def calculate_3_points(diameter, thickness_12, thickness_3, thickness_9):
    # Use diameter of outer circle (cavity) and wall thickness in 3 places (@12, @3, @9) to calculate 3 points of inner circle (core)
    # Wall thickness @12, @3, @9 are chosen to make calculations more simple

    radius = diameter / 2
    point_12 = (0, radius + thickness_12)
    point_3 = (radius + thickness_3, 0)
    point_9 = (-1 * (radius + thickness_9), 0)

    return point_12, point_3, point_9

def circ_CAV(x, y, dia):
    # Draw circle to show CAVITY

    dia = dia * global_scale

    draw.ellipse(
        (x-dia / 2, y-dia / 2, x+dia / 2, y+dia / 2),
        outline = 'red',
        width = 2
        )

def circ_COR(x, y, dia):
    # Draw circle to show CORE

    dia = dia * global_scale
    draw.ellipse(
        (x-dia / 2, y-dia / 2, x+dia / 2, y+dia / 2),
        outline = 'green',
        width = 5
        )

def cav_nr(x, y, nr):
    # Draw cavity number (if cavity number is single digit, add '0' in front of that digit)

    if len(str(nr))<2:
        nr = '0' + str(nr)
    else:
        nr = str(nr)

    font = ImageFont.truetype("arial.ttf", 8 * global_scale)

    draw.text(
        (x, y - 5 * global_scale),
        nr,
        fill = 'white',
        stroke_width = 1,
        stroke_fill = 'black',
        font = font,
        anchor = 'mm'
        )

def adjustment_text(x, y, txt1, txt2):
    # Print information: distance between centers of core and cavity circles

    # Add empty space in front of positive value to make both strings the same length
    if txt1[0] != '-':
        txt1 = ' ' + txt1
    if txt2[0] != '-':
        txt2 = ' ' + txt2

    # Add 'x: ' and 'y: ' in fron of values, add '0' if it is needed to make length of both strings equal
    if len(txt1) > len(txt2):
        x_adjusted = 'x: ' + txt1
        y_adjusted = 'y: ' + txt2 + '0'
    elif len(txt1) < len(txt2):
        x_adjusted = 'x: ' + txt1 + '0'
        y_adjusted = 'y: ' + txt2
    else:
        x_adjusted = 'x: ' + txt1
        y_adjusted = 'y: ' + txt2

    font = ImageFont.truetype("arial.ttf", 3 * global_scale)

    draw.text(
        (x, y + 2.5 * global_scale),
        x_adjusted,
        fill='black',
        font=font,
        anchor='mm'
        )

    draw.text(
        (x, y + 5.5 * global_scale),
        y_adjusted,
        fill='black',
        font=font,
        anchor='mm'
        )

def adjust_string_length(x):
    # Helper function to add trailing 0 in front of a single digit number

    if len(x) < 2:
        x = '0' + x
    return(x)

def draw_header():

    report_file = excel_files_in_directory()
    now = datetime.now()
    report_date = str(now.year) + '.' + adjust_string_length(str(now.month)) + '.' + adjust_string_length(str(now.day))
    report_time = adjust_string_length(str(now.hour)) + ':' + adjust_string_length(str(now.minute))

    text_color = 'black'

    draw.line(
        (0 + global_scale * 5, spacing / 2, width - global_scale * 5, spacing / 2),
        fill = 'black',
        width = 3
        )

    font = ImageFont.truetype("arial.ttf", 5 * global_scale)

    draw.text(
        (0 + global_scale * 5, spacing / 7),
        ("Image generated from file: '" + report_file + "'"),
        fill = text_color,
        font = font,
        anchor = 'lt'
        )
    draw.text(
        (0 + global_scale * 5, spacing / 3.2),
        (report_date + '   ' + report_time),
        fill = text_color,
        font = font,
        anchor = 'lt'
        )

    #print("Image generated from file: '" + report_file + "'")
    #print('Image generated at: ' + report_date)

def draw_all():

    draw_header()

    count = 0

    x = spacing / 2
    y = spacing 

    for coll in range(colls):
        for row in range(rows):
            # If there is no data for that cavity, go to next cavity
            if type(data[count][0]) != float:
                y = y + spacing
                count = count + 1
            else:
                # Draw cavity circle in grid (distances between cavity cirsles are uniform)
                circ_CAV(
                    x,               # x coordinate
                    y,               # y coordinate
                    data[count][0]   # diameter of circle
                    )

                # Calculate, where to draw core circle
                points = calculate_3_points(data[count][1], data[count][2], data[count][3], data[count][4])
                center_adjustment = calculate_circle_center(points[0], points[1], points[2])

                # Draw core circle
                circ_COR(
                    x + center_adjustment[0] * 10 * global_scale,     # x coordinate
                    y + center_adjustment[1] * 10 * global_scale,     # y coordinate
                    data[count][1] * 0.85 #* global_scale               # diameter of circle
                    )

                # Draw cavity number
                cav_nr(x, y, count+1)

                # Draw information about difference between cavity and core centre positions
                adjustment_text(x, y, str(round(center_adjustment[0], 3)), str(round(center_adjustment[1], 3)))

                y = y + spacing
                count = count + 1

        x = x + spacing
        y = spacing

# Read data from excel file:
data_file = excel_files_in_directory()

# Based on mold size define size and layout of image
rows_start, rows_finish, number_of_cavities = get_range()[0], get_range()[1], get_range()[2]

if number_of_cavities == 96:
    print('Number of cavities is 96')
    colls, rows = 8, 12
elif number_of_cavities == 72:
    print('Number of cavities is 72')
    colls, rows = 6, 12
else:
    print('Unexpected number of cavities (must be 96 or 72)')

spacing = 37.5 * global_scale
width, height = int(colls * spacing), int(rows * spacing + spacing / 2)

# Read data from excel file
data = data_list(data_file, rows_start, rows_finish)

# Create image
img = Image.new("RGB", (width, height), "white")
draw = ImageDraw.Draw(img)

# Draw all the information on image
draw_all()

# Show and save image (saves in the same dir as script file)
img.show()
#img.save("output.png", format="PNG")

