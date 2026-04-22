import os
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from datetime import datetime

# Script for visualisation of TT (Top panel Thickness) measurement data for CIC-1810 closures

# What this script does:
# 1. Find CIC-1810 excel file in the same folder as this script
# 2. Read TT measurement data from 'TT measurements' sheet (12 positions per cavity)
# 3. Draw polar diagram for each cavity: 12 coloured sectors showing thickness deviation
#    Colours: white=nominal, blue=below nominal (in tolerance), orange=above nominal (in tolerance),
#             red=above tolerance limit, purple=below tolerance limit

global_scale = 10


def excel_files_in_directory():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    all_files = os.listdir(script_dir)
    files_to_return = [f for f in all_files if f.endswith('.xlsx') and 'CIC-1810' in f]

    if len(files_to_return) > 1:
        print('There are several CIC-1810 excel files in this folder!')
        return None
    if len(files_to_return) == 0:
        print('There are no CIC-1810 excel files in this folder!')
        return None
    return files_to_return[0]


def get_range():
    data_file = excel_files_in_directory()
    wb = load_workbook(filename=data_file)
    ws = wb['TT measurements']

    position = 32
    count = 0
    reference_type = type(ws['B32'].value)

    while True:
        cell_value = ws['B' + str(position)].value
        if type(cell_value) != reference_type:
            break
        position += 1
        count += 1

    return 32, position - 1, count


def data_list(data_file, start, stop):
    wb = load_workbook(filename=data_file)
    ws = wb['TT measurements']

    nominal = ws['C20'].value
    utv    = ws['C21'].value
    ltv    = ws['C22'].value

    columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
    result = []

    for row_num in range(start, stop + 1):
        measurements = [ws[col + str(row_num)].value for col in columns]
        result.append(measurements)

    return result, nominal, utv, ltv


def get_sector_color(value, nominal, upper_limit, lower_limit):
    if value is None:
        return (200, 200, 200)
    if value > upper_limit:
        return (220, 0, 0)
    if value < lower_limit:
        return (128, 0, 128)
    if abs(value - nominal) < 0.001:
        return (255, 255, 255)
    if value > nominal:
        return (255, 165, 0)
    return (0, 100, 220)


def draw_polar(x, y, measurements, cav_num):
    radius = 14 * global_scale
    angle_step = 30

    for i, value in enumerate(measurements):
        color = get_sector_color(value, nominal, upper_limit, lower_limit)
        start_angle = -90 + i * angle_step
        end_angle   = start_angle + angle_step

        draw.pieslice(
            (x - radius, y - radius, x + radius, y + radius),
            start=start_angle,
            end=end_angle,
            fill=color,
            outline='black',
            width=1
        )

    font = ImageFont.truetype("arial.ttf", 4 * global_scale)
    nr = str(cav_num).zfill(2)
    draw.text((x, y), nr, fill='black', font=font, anchor='mm')


def adjust_string_length(x):
    if len(x) < 2:
        x = '0' + x
    return x


def draw_header():
    report_file = excel_files_in_directory()
    now = datetime.now()
    report_date = str(now.year) + '.' + adjust_string_length(str(now.month)) + '.' + adjust_string_length(str(now.day))
    report_time = adjust_string_length(str(now.hour)) + ':' + adjust_string_length(str(now.minute))

    draw.line(
        (5 * global_scale, spacing / 2, width - 5 * global_scale, spacing / 2),
        fill='black',
        width=3
    )

    font_large = ImageFont.truetype("arial.ttf", 5 * global_scale)
    draw.text(
        (5 * global_scale, spacing / 7),
        "Image generated from file: '" + report_file + "'",
        fill='black',
        font=font_large,
        anchor='lt'
    )
    draw.text(
        (5 * global_scale, spacing / 3.2),
        report_date + '   ' + report_time,
        fill='black',
        font=font_large,
        anchor='lt'
    )

    legend_x = width - 70 * global_scale
    legend_y = int(spacing / 7)
    font_legend = ImageFont.truetype("arial.ttf", 3 * global_scale)

    legend_items = [
        ((220, 0, 0),     '> ' + str(round(upper_limit, 2)) + ' (above tolerance)'),
        ((255, 165, 0),   str(nominal) + ' - ' + str(round(upper_limit, 2)) + ' (above nominal)'),
        ((255, 255, 255), '= ' + str(nominal) + ' (nominal)'),
        ((0, 100, 220),   str(round(lower_limit, 2)) + ' - ' + str(nominal) + ' (below nominal)'),
        ((128, 0, 128),   '< ' + str(round(lower_limit, 2)) + ' (below tolerance)'),
    ]

    for idx, (color, label) in enumerate(legend_items):
        bx = legend_x
        by = legend_y + idx * 5 * global_scale
        draw.rectangle(
            (bx, by, bx + 4 * global_scale, by + 4 * global_scale),
            fill=color,
            outline='black',
            width=1
        )
        draw.text(
            (bx + 5 * global_scale, by + 2 * global_scale),
            label,
            fill='black',
            font=font_legend,
            anchor='lm'
        )


def draw_all():
    draw_header()

    count = 0
    x = spacing / 2
    y = spacing

    for col in range(colls):
        for row in range(rows):
            if count < len(data):
                draw_polar(x, y, data[count], count + 1)
                count += 1
            y += spacing
        x += spacing
        y = spacing


# Read data
data_file = excel_files_in_directory()
rows_start, rows_finish, number_of_cavities = get_range()

if number_of_cavities == 96:
    print('Number of cavities is 96')
    colls, rows = 8, 12
elif number_of_cavities == 72:
    print('Number of cavities is 72')
    colls, rows = 6, 12
else:
    print('Unexpected number of cavities (must be 96 or 72)')

spacing = 37.5 * global_scale
width  = int(colls * spacing)
height = int(rows * spacing + spacing / 2)

data, nominal, utv_raw, ltv_raw = data_list(data_file, rows_start, rows_finish)
upper_limit = nominal + utv_raw
lower_limit = nominal + ltv_raw

# Create image
img  = Image.new("RGB", (width, height), "white")
draw = ImageDraw.Draw(img)

draw_all()

img.show()
# img.save("output_TT.png", format="PNG")
